# librairies
import pandas as pd
import os
import re
import tempfile
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from parser_TT140_MasterCard import *
import streamlit as st
import io
import numpy as np
import zipfile


# Define the function to read CSV files with delimiters
def read_csv_with_delimiters(file_path, default_columns=None, default_delimiter=','):
    """
    Lire un fichier CSV avec delimiteurs , ou ; ou espace
    
    Les paramètres:
        Chemin de fichier (str): Le chemin du fichier CSV.
        Le delimiteur par défaut (str): Le delimiteur par défaut à utiliser si le fichier est vide.
        
    Retourne:
        pd.DataFrame: Le dataframe avec le contenu CSV.
    """
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline()

        # Détecter le délimiteur
        if ';' in first_line:
            delimiter = ';'
        elif ',' in first_line:
            delimiter = ','
        elif ' ' in first_line:
            delimiter = '\s+'  # regex for one or more spaces
        else:
            delimiter = default_delimiter
    except FileNotFoundError:
        delimiter = default_delimiter

    try:
        df = pd.read_csv(file_path, sep=delimiter, engine='python')
    except pd.errors.EmptyDataError:
        df = pd.DataFrame(columns=default_columns)
    
    return df


# Function to save uploaded file to a temporary location
def save_uploaded_file(uploaded_file):
    with tempfile.NamedTemporaryFile(delete=False) as temp_file:
        temp_file.write(uploaded_file.read())
        return temp_file.name
    

default_columns_cybersource = ['NBRE_TRANSACTION', 'MONTANT_TOTAL', 'CUR', 'FILIALE', 'RESEAU', 'TYPE_TRANSACTION']
default_columns_saisie_manuelle = ['NBRE_TRANSACTION', 'MONTANT_TOTAL', 'CUR', 'FILIALE', 'RESEAU']
default_columns_pos = ['FILIALE', 'RESEAU', 'TYPE_TRANSACTION', 'DATE_TRAI', 'CUR', 'NBRE_TRANSACTION', 'MONTANT_TOTAL']



def reading_cybersource(cybersource_file):
    # Lire le fichier cybersource
    if os.path.exists(cybersource_file):
        df_cybersource = read_csv_with_delimiters(cybersource_file, default_columns_cybersource)

        # nettoyer les espaces à partir des noms de colonnes
        df_cybersource.columns = df_cybersource.columns.str.strip()
        
        # affecter aux transactions Cybersource le type ACHAT pour merger après avec les transactions POS de type ACHAT
        df_cybersource['TYPE_TRANSACTION'] = 'ACHAT'
        df_cybersource = df_cybersource.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        #df_cybersource['RESEAU'] = df_cybersource['RESEAU'].astype(str)
        #df_cybersource['FILIALE'] = df_cybersource['FILIALE'].astype(str)
        #df_cybersource['CUR'] = df_cybersource['CUR'].astype(str)
        return df_cybersource
    else:
        df_cybersource = pd.DataFrame(columns=default_columns_cybersource)
        print("The Cybersource file does not exist at the specified path.")


# Read Saisie Manuelle file
def reading_saisie_manuelle(saisie_manuelle_file):
    # Lire le fichier saisie manuelle
    if os.path.exists(saisie_manuelle_file):
        df_sai_manuelle = read_csv_with_delimiters(saisie_manuelle_file, default_columns_saisie_manuelle)

        # nettoyer les espaces à partir des noms de colonnes
        df_sai_manuelle.columns = df_sai_manuelle.columns.str.strip()

        # affecter aux transactions Cybersource le type ACHAT pour merger après avec les transactions POS de type ACHAT
        df_sai_manuelle['TYPE_TRANSACTION'] = 'ACHAT'
        df_sai_manuelle = df_sai_manuelle.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        return df_sai_manuelle
    else:
        df_sai_manuelle = pd.DataFrame(columns=default_columns_saisie_manuelle)
        print("The Saisie Manuelle file does not exist at the specified path.")


# Read POS file
def reading_pos(pos_file):
    # Lire le fichier POS
    if os.path.exists(pos_file):
        df_pos = read_csv_with_delimiters(pos_file, default_columns_pos)

        # nettoyer les espaces à partir des noms de colonnes
        df_pos.columns = df_pos.columns.str.strip()
        df_pos = df_pos.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

        # renommer la colonne BANQUE en colonne FILIALE pour normaliser le dataframe avec les autres dataframes des autres sources
        df_pos.rename(columns={'BANQUE': 'FILIALE'}, inplace=True)
        return df_pos
    else:
        df_pos = pd.DataFrame(columns=default_columns_pos)
        print("The POS file does not exist at the specified path.")


def filtering_sources(df_cybersource, df_sai_manuelle, df_pos):
    # Filter chaque source pour recevoir uniquement les transactions de type VISA INTERNATIONAL
    filtered_cybersource_df = df_cybersource[df_cybersource['RESEAU'] == 'VISA INTERNATIONAL']
    filtered_saisie_manuelle_df = df_sai_manuelle[df_sai_manuelle['RESEAU'] == 'VISA INTERNATIONAL']
    filtered_pos_df = df_pos[(df_pos['RESEAU'] == 'VISA INTERNATIONAL') & 
                            (~df_pos['TYPE_TRANSACTION'].str.endswith('_MDS'))]
    return filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df


def validate_file_name_and_date(file_name, source, date_to_validate=None):
    """
    Validate the file name based on the source, the required pattern,
    and optionally validate the date extracted from the file name.

    Parameters:
        file_name (str): The name of the uploaded file.
        source (str): The source type (CYBERSOURCE, POS, or SAIS_MANU).
        date_to_validate (str): The date to validate against the one in the file name. (Optional)

    Returns:
        bool: True if the file name is valid, False otherwise.

    Raises:
        ValueError: If the file name is invalid or if the date does not match the expected pattern.
    """
    pattern = f"^TRANSACTION_{source}_TRAITE_SG_\\d{{2}}-\\d{{2}}-\\d{{2}}_\\d{{6}}\\.CSV$"
    if not re.match(pattern, file_name):
        raise ValueError(f"Invalid file name: {file_name}. Expected pattern: TRANSACTION_{source}_TRAITE_SG_YY-MM-DD_HHMMSS.CSV")

    # Extract the date from the file name
    date_match = re.search(r"\d{2}-\d{2}-\d{2}", file_name)
    if date_match:
        extracted_date = date_match.group(0)
        if date_to_validate:
            if extracted_date != date_to_validate:
                raise ValueError(f"Date extracted from the file name ({extracted_date}) does not match the provided date ({date_to_validate}).")
    else:
        raise ValueError("Date not found in the file name.")

    return True


# Fonction qui convertit un fichier excel en csv en dataframe, pour le cas du fichier des rejets recyclées
def excel_to_csv_to_df(excel_file_path, sheet_name=0):
    """
    Convertit un fichier xlsx en csv en dataframe.

    Parameters:
        excel_file_path (str): Chemin pour le fichier excel.
        nom de la feuille excel (str or int): Nom ou index de la feuille excel (Par défaut la première feuille).

    Returns:
        pd.DataFrame: Dataframe contenant les donnèes de csv.
    """
    try:
        # Lire le fichier excel
        df = pd.read_excel(excel_file_path,engine='openpyxl', sheet_name=sheet_name, header=0)

        # Définir le chemin du fichier csv
        csv_file_path = excel_file_path.replace('.xlsx', '.csv')

        # Sauvegarder le dataframe en csv
        df.to_csv(csv_file_path, index=False)

        # Lire le csv en dataframe
        df_csv = read_csv_with_delimiters(csv_file_path)
        return df_csv

    except PermissionError as p_error:
        print(f"Permission error: {p_error}. Please check your file permissions.")
    except Exception as e:
        print(f"An error occurred: {e}")
    return None


def standardize_date_format(date_column, desired_format='%Y-%m-%d'):
    """
    Standardize the date format in a given column.

    Parameters:
        date_column (pd.Series): The column containing dates to standardize.
        desired_format (str): The desired date format (default is '%Y-%m-%d').

    Returns:
        pd.Series: The column with dates in the standardized format.
    """
    # Convert all dates to datetime objects
    date_column = pd.to_datetime(date_column , dayfirst=False  , yearfirst=True)
    # Format all datetime objects to the desired format
    date_column = date_column.dt.strftime(desired_format )

    return date_column

def merging_sources_without_recycled(filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df, file_path, content, zip_file_path, zip_reject_path):

    # Ensure 'TYPE_TRANSACTION' column exists in each DataFrame, if not, create it with default empty values
    if 'TYPE_TRANSACTION' not in filtered_cybersource_df.columns:
        filtered_cybersource_df['TYPE_TRANSACTION'] = ''
    
    if 'TYPE_TRANSACTION' not in filtered_saisie_manuelle_df.columns:
        filtered_saisie_manuelle_df['TYPE_TRANSACTION'] = ''
    
    if 'TYPE_TRANSACTION' not in filtered_pos_df.columns:
        filtered_pos_df['TYPE_TRANSACTION'] = ''

    # Filter cybersource_df to include only rows with TYPE_TRANSACTION == 'ACHAT'
    filtered_cybersource_df = filtered_cybersource_df[filtered_cybersource_df['TYPE_TRANSACTION'] == 'ACHAT']

    # Fusionner avec le dataframe POS avec le dataframe SAISIE MANUELLE en se basant sur les colonnes précisées
    result_df = pd.merge(
        filtered_pos_df,
        filtered_saisie_manuelle_df,
        on=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION'],
        suffixes=('_pos', '_saisie'),
        how='outer'  # Utiliser fusionnage de type outer pour garder toutes les lignes
    )

    # Fusionner avec les dataframes fusionnés (POS et SAISIE MANUELLE) avec le dataframe CYBERSOURCE filtré
    result_df = pd.merge(
        result_df,
        filtered_cybersource_df,
        on=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION'],
        suffixes=('_merged', '_cybersource'),
        how='outer'  # Utiliser fusionnage de type outer pour garder toutes les lignes
    )

    # Remplacer les valeurs nulles avec 0, et sommer les nombres de transactions
    result_df['NBRE_TRANSACTION'] = (result_df['NBRE_TRANSACTION_pos'].fillna(0) +
                                    result_df['NBRE_TRANSACTION_saisie'].fillna(0) +
                                    result_df['NBRE_TRANSACTION'].fillna(0))

    # Convertir le 'NBRE_TRANSACTION' en entier
    result_df['NBRE_TRANSACTION'] = result_df['NBRE_TRANSACTION'].astype(int)

    # Sommer les montants de toutes les sources
    result_df['MONTANT_TOTAL'] = (result_df['MONTANT_TOTAL_pos'].fillna(0) +
                                    result_df['MONTANT_TOTAL_saisie'].fillna(0) +
                                    result_df['MONTANT_TOTAL'].fillna(0))

    # Supprimer les colonnes inecessaires utilisées en fusionnage
    result_df.drop(['NBRE_TRANSACTION_pos', 'NBRE_TRANSACTION_saisie', 'MONTANT_TOTAL_pos', 'MONTANT_TOTAL_saisie'], axis=1, inplace=True)

    # Supprimer les lignes dupliquées
    result_df.drop_duplicates(subset=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION', 'DATE_TRAI'], inplace=True)

    total_nbre_transactions = result_df['NBRE_TRANSACTION'].sum()

    result_df = result_df.reset_index(drop=True)




    # Définir les colonnes du dataframe du résultat de réconciliation
    new_columns = [
        'FILIALE', 'Réseau', 'Type', 'Date', 'Devise', 'NbreTotaleDeTransactions',
        'Montant Total de Transactions', 'Rapprochement', 'Nbre Total de Rejets',
        'Montant de Rejets', 'Nbre de Transactions (Couverture)', 
        'Montant de Transactions (Couverture)'
    ]

    # Mapper entre les noms de colonnes (merged_df et le dataframe résultat)
    column_mapping = {
        'FILIALE': 'FILIALE',
        'RESEAU': 'Réseau',
        'TYPE_TRANSACTION': 'Type',
        'DATE_TRAI': 'Date',
        'CUR': 'Devise',
        'NBRE_TRANSACTION': 'NbreTotaleDeTransactions',
        'MONTANT_TOTAL': 'Montant Total de Transactions'
    }

    # Renommer les colonnes
    result_df.rename(columns=column_mapping, inplace=True)

    # Remplacer les colonnes vides avec colonnes par défaut
    for column in set(new_columns) - set(result_df.columns):
        result_df[column] = ''

    # Réordonner les colonnes
    result_df = result_df[new_columns]

    #print(result_df)

    filename = os.path.basename(file_path)
    bin_number = filename.split('_')[0]  # Extraire BIN des fichiers


     # pour ouvrir en cas de présence de dossier
     # with open(file_path, 'r') as file:
        #content = file.read()
        
    cleaned_content = re.sub(
    r'(ISSUER TRANSACTIONS.*?(?:\*\*\*  END OF VSS-120 REPORT|FINAL SETTLEMENT NET AMOUNT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-130.*?(?:\*\*\*  END OF VSS-130 REPORT)|'
    r'REPORT ID:\s+VSS-140.*?(?:\*\*\*  END OF VSS-140 REPORT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-115.*?(?:\*\*\*  END OF VSS-115 REPORT)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+DISPUTE FIN.*?TOTAL PURCHASE\s+([\d,]+)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+ DISPUTE RESP FIN.*?TOTAL PURCHASE\s+([\d,]+))', 
    '', 
    content, 
    flags=re.DOTALL
)

    # Chercher la section d'extraction commençant de "ACQUIRER TRANSACTIONS" jusqu'à "END OF VSS-120 REPORT"
    acquirer_section = re.search(r'ACQUIRER TRANSACTIONS(.*)(?=\*\*\*  END OF VSS-120 REPORT)', cleaned_content, re.DOTALL)

    if acquirer_section:
        acquirer_content = acquirer_section.group(1).strip()  # Extraire et nettoyer le contenu du bloc

        # Extraire le nombre de transactions TOTAL ORIGINAL SALE
        total_original_sale = re.search(r'TOTAL PURCHASE\s+([\d,]+)', acquirer_content)
        if total_original_sale:
            transaction_count = total_original_sale.group(1).replace(',', '')
        else:
            transaction_count = None

        # Extraire le nombre de transactions MANUAL CASH (CASH ADVANCE)
        manual_cash_strict = re.search(r'^\s*MANUAL CASH\s+(\d+)', acquirer_content, re.MULTILINE)

        total_purchase_mga = re.search(r'CLEARING CURRENCY:\s+MGA.*?TOTAL PURCHASE\s+([\d,]+)', acquirer_content, re.DOTALL)

        # Extract le nombre de transactions TOTAL PURCHASE (ACHAT) pour MGA CLEARING CURRENCY EUR
        nbr_transaction_eur = re.search(r'CLEARING CURRENCY:\s+EUR.*?ORIGINAL SALE\s+(\d+)', acquirer_content, re.DOTALL)
        
        if nbr_transaction_eur:
            transaction_count_EUR = nbr_transaction_eur.group(1).replace(',', '')
        else:
            transaction_count_EUR = None

        if total_purchase_mga:
            transaction_count_MGA = total_purchase_mga.group(1).replace(',', '')
        else:
            transaction_count_MGA = None

        # Extraire le nombre de transactions MERCHANDISE CREDIT (CREDIT VOUCHER)
        merchandise_credit_section = re.search(r'TOTAL MERCHANDISE CREDIT\s+(\d+)', acquirer_content)

        # Extraire le nombre de rejets
        total_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if total_rejects:
            transaction_rejects_count = total_rejects.group(1) #Prendre les elements de la première position comme élement d'extraction
        else:
            transaction_rejects_count = "None"
        
        xof_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if xof_rejects:
            xof_rejects_count = xof_rejects.group(1) #Prendre les elements de la deuxième position comme élement d'extraction (XOFS)
        else:
            xof_rejects_count = "None"

        # Créer un dictionnaire pour stocker les résultats
        transaction_data = {
            "file": file_path,
            "Numero de BIN": bin_number,
            "Nbr de transactions": transaction_count,
            "CASH ADVANCE": None,
            "Nbr de transactions (CLEARING CURR EUR POUR MDG)": None,
            "CREDIT VOUCHER": None,
            "total de rejets" : transaction_rejects_count,
            "total de xof rejetées" : xof_rejects_count
        }

        # Ajouter le nombre de transactions MANUAL CASH (CASH ADVANCE)
        if manual_cash_strict:
            manual_cash_value = manual_cash_strict.group(1)
            if manual_cash_value != "0":  # Ignorer si MANUAL CASH est 0
                transaction_data["CASH ADVANCE"] = manual_cash_value

        # Ajouter le nombre de transactions pour TOTAL PURCHASE (CLEARING CURRENCY EUR) seulement si le BIN est 489316
        if bin_number == "489316":
            transaction_data["Nbr de transactions (CLEARING CURR EUR POUR MDG)"] = transaction_count_EUR
            transaction_data["Nbr de transactions"] = transaction_count_MGA

        # Ajouter le nombre de transactions pour TOTAL MERCHANDISE CREDIT (CREDIT VOUCHER) si ça existe
        if merchandise_credit_section:
            merchandise_credit_count = merchandise_credit_section.group(1)
            transaction_data["CREDIT VOUCHER"] = merchandise_credit_count

    else:
        return {"file": file_path, "error": "ACQUIRER TRANSACTIONS section not found."}
    
    # appliquer l'extraction sur les fichiers à travers le dossier des rapports de SETTLEMENT
    transaction_list = []
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        # Lister tous les fichiers dans le ZIP
        for file_name in zip_ref.namelist():
            if file_name.endswith('.TXT'):
                # Extraire le fichier en mémoire
                with zip_ref.open(file_name) as file:
                    content = file.read().decode('utf-8')
                    # Appliquer l'extraction des données sur le fichier
                    transaction_data = extract_transaction_data(file_name, content)
                    transaction_list.append(transaction_data)
        
    
    # boucler sur le dictionnaire pour afficher les transactions pour chaque filiale (en se basant sur le BIN)
    for transaction in transaction_list:
        print(transaction)


    # Initialiser la mapping des FILIALES à BINs
    visa_banks_bin = {
        'SG - COTE D IVOIRE': '463741',
        'SG - BENIN': '404927',
        'SG - BURKINA FASO': '410282',
        'SG - CAMEROUN': '439972',
        'SG - GUINEE EQUATORIALE': '410655',
        'SG - MADAGASCAR': '489316',
        'SG - SENEGAL': '441358',
        'SG - TCHAD': '458250',
        'SG - CONGO': '464012',
        'SG - GUINEE CONAKRY': '486059'
    }


    # Créer un dictionnaire à partir de la liste des transactions extraites
    generated_transactions = {trans['Numero de BIN']: trans for trans in transaction_list if 'error' not in trans}

    # Pre-calculer (faire la somme) les lignes ayant comme TYPES :  ACHAT et CREDIT VOUCHER groupées par filiale
    grouped_filiale = result_df.groupby('FILIALE').agg(
        achat_transactions=pd.NamedAgg(column='NbreTotaleDeTransactions', aggfunc=lambda x: x[result_df['Type'] == 'ACHAT'].sum()),
        cv_transactions=pd.NamedAgg(column='NbreTotaleDeTransactions', aggfunc=lambda x: x[result_df['Type'].isna() | (result_df['Type'] == '')].sum())
    ).reset_index()

    def safe_float(value):
        """Convert a value to float, handling 'None' and NoneType."""
        if value in (None, 'None'):
            return 0
        try:
            return float(value)
        except ValueError:
            return 0  # Default 0 is la conversion est failed

    def compare_transactions(row, transaction_data, grouped_data):
        filiale = row['FILIALE']
        transaction_type = row['Type']

          
        # Trouver le BIN correspondant par FILIALE
        bin_number = visa_banks_bin.get(filiale)
        if not bin_number:
            return 'Aucun BIN trouvé'
        
        # Extraire les donnèes de transactions par BIN
        trans_data = transaction_data.get(bin_number)
        if not trans_data:
            return 'Aucune donnèe pour BIN'
        
        # Extraire les informations de transactions du dictionnaire
        transaction_count = safe_float(trans_data.get('Nbr de transactions'))
        total_count = safe_float(trans_data.get('Nbr de transactions (CLEARING CURR EUR POUR MDG)'))
        manual_cash = safe_float(trans_data.get('CASH ADVANCE'))
        merchandise_credit = safe_float(trans_data.get('CREDIT VOUCHER'))
        rejects = safe_float(trans_data.get('total de rejets'))
        #xof_rejects = safe_float(trans_data.get('total de xof rejetées'))

        # Get sommes pré-calculées à travers les donnèes groupées
        filiale_data = grouped_data[grouped_data['FILIALE'] == filiale]
        if filiale_data.empty:
            return 'Filiale inconnue'

        achat_transactions = safe_float(filiale_data['achat_transactions'].values[0])
        cv_transactions = safe_float(filiale_data['cv_transactions'].values[0])

        #print(f"FILIALE: {filiale}")
        #print(f"ACHAT transactions (FILIALE): {achat_transactions}")
        #print(f"CREDIT VOUCHER transactions (FILIALE): {cv_transactions}")
        #print(f"Transaction count from settlement report: {transaction_count}")
        #print(f"Rejects from settlement report: {rejects}")
        #print(f"MERCHANDISE CREDIT from settlement report: {merchandise_credit}")

        #Le cas ou on a pas de rejets
        if rejects == 0:
            if transaction_type == 'ACHAT':
                filiale = row['FILIALE']

                # Checker si on a CREDIT VOUCHER (vide/NaN 'Type') pour la meme FILIALE
                cv_rows = result_df[(result_df['FILIALE'] == filiale) & (pd.isna(result_df['Type']) | (result_df['Type'] == ''))]

                # If no CREDIT VOUCHER exists for this FILIALE, apply additional checks
                if cv_rows.empty:
                
                    # Apply checks on transaction counts
                    if row['NbreTotaleDeTransactions'] == transaction_count + merchandise_credit or row['NbreTotaleDeTransactions'] == total_count:
                    
                        return 'ok'
                    else:
                    
                        return 'not ok (avec tr.(s) rejetée(s) a extraire)'

                else:
    
                    if row['NbreTotaleDeTransactions'] == transaction_count or row['NbreTotaleDeTransactions'] == total_count:
                    
                        return 'ok'
                    else:
                    
                        return 'not ok (avec tr.(s) rejetée(s) a extraire)'
                
                
                
            if transaction_type == 'ACHAT' and transaction_type == '' or  pd.isna(transaction_type):

                # grouper les lignes par filiale
                filiale = row['FILIALE']

                # lignes avec meme filiale pour ACHAT et CREDIT VOUCHER
                relevant_rows = result_df[(result_df['FILIALE'] == filiale) &
                                    ((result_df['Type'] == 'ACHAT') | pd.isna(result_df['Type']) | (result_df['Type'] == ''))]
                

                # Sommer 'NbreTotaleDeTransactions' pour les lignes
                total_transactions = relevant_rows['NbreTotaleDeTransactions'].sum()

                # formule de comparaison
                formule = total_transactions == transaction_count + merchandise_credit
                print(f"Formule sans calcul appliquée oui ou non ?: {formule_calcul}\n")
                if formule:
                    return 'ok'
                else:
                    print(total_transactions)
                    print("crd v + Trans c")
                    print(transaction_count + merchandise_credit)
                    return 'not ok (avec tr.(s) rejetée(s) a extraire)'
                
            elif transaction_type == 'CASH ADVANCE':
                if row['NbreTotaleDeTransactions'] == manual_cash:
                    return 'ok'
                else:
                    return 'not ok (avec tr.(s) rejetée(s) a extraire)'
            else:
                return 'ok'
            
        # Le cas ou on a des rejets
        elif rejects != 0:
            try:
                formule_calcul = (achat_transactions + cv_transactions) == (transaction_count - rejects + merchandise_credit)
                print(f"Formule de rejets calcul appliquée oui ou non ?: {formule_calcul}\n")

                if formule_calcul  and transaction_type == 'ACHAT':
                    return 'not ok'
                elif formule_calcul and (transaction_type == '' or  pd.isna(transaction_type)):
                    return 'ok'
                elif not formule_calcul and (transaction_type == '' or  pd.isna(transaction_type)): 
                    return 'ok'
                elif not formule_calcul and transaction_type == 'ACHAT':
                    return 'not ok (avec tr.(s) rejetée(s) a extraire)'

            except ValueError:
                return 'not ok'  # Gérer les erreurs de conversion

        else:
            return 'Type inconnu'
        
    def update_rejects(data, transaction_data):
        for idx, row in data.iterrows():
            # Ensure 'Rapprochement' is not None and is a string
            rapprochement = row.get('Rapprochement')
            #if not isinstance(rapprochement, str):
                #continue

            if 'not ok (avec tr.(s) rejetée(s) a extraire)' in rapprochement:
                # Fill only 'Nbre Total de Rejets', leave 'Montants de rejets' empty
                filiale = row['FILIALE']
                bin_number = visa_banks_bin.get(filiale)
                trans_data = transaction_data.get(bin_number)
                if trans_data:
                    data.at[idx, 'Nbre Total de Rejets'] = trans_data.get('total de rejets', 0)
                    data.at[idx, 'Montant de Rejets'] = trans_data.get('total de xof rejetées', 0)
            elif 'not ok' in rapprochement:
                # Fill both 'Nbre Total de Rejets' and 'Montants de rejets'
                filiale = row['FILIALE']
                bin_number = visa_banks_bin.get(filiale)
                trans_data = transaction_data.get(bin_number)
                if trans_data:
                    data.at[idx, 'Nbre Total de Rejets'] = trans_data.get('total de rejets', 0)
                    data.at[idx, 'Montant de Rejets'] = trans_data.get('total de xof rejetées', 0)

    

        # Fonction pour extraire les rejets des rapports SETTLEMENT de VISA
    def extract_rejects_data(file_path, content):
        filename = os.path.basename(file_path)
        bin_number = filename.split('_')[0]  # Extraire le BIN du nom de fichier

        #with open(file_path, 'r') as file:
            #content = file.read()


        # Nettoyer le contenu pour supprimer les sections non pertinentes
        cleaned_content = re.sub(
        r'(ISSUER TRANSACTIONS.*?(?:\*\*\*  END OF VSS-120 REPORT|FINAL SETTLEMENT NET AMOUNT|TOTAL MERCHANDISE CREDIT)|'
        r'REPORT ID:\s+VSS-130.*?(?:\*\*\*  END OF VSS-130 REPORT)|'
        r'REPORT ID:\s+VSS-140.*?(?:\*\*\*  END OF VSS-140 REPORT|TOTAL MERCHANDISE CREDIT)|'
        r'REPORT ID:\s+VSS-115.*?(?:\*\*\*  END OF VSS-115 REPORT)|'
        r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+DISPUTE FIN.*?TOTAL PURCHASE\s+([\d,]+)|'
        r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+ DISPUTE RESP FIN.*?TOTAL PURCHASE\s+([\d,]+))', 
        '', 
        content, 
        flags=re.DOTALL
    )
        
        # Chercher la section d'extraction commençant de "ACQUIRER TRANSACTIONS" jusqu'à "END OF VSS-120 REPORT"
        acquirer_section = re.search(r'ACQUIRER TRANSACTIONS(.*)(?=\*\*\*  END OF VSS-120 REPORT)', cleaned_content, re.DOTALL)

        if acquirer_section:
            acquirer_content = acquirer_section.group(1).strip()  # Extraire et nettoyer le contenu du bloc

            # Extraire le nombre de rejets
            total_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
            if total_rejects:
                transaction_rejects_count = total_rejects.group(1)  # Prendre les éléments de la première position comme élément d'extraction
            else:
                transaction_rejects_count = "None"
            
            # Extraire le nombre de rejets xof
            xof_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
            if xof_rejects:
                xof_rejects_count = xof_rejects.group(1)  # Prendre les éléments de la deuxième position comme élément d'extraction (XOFS)
            else:
                xof_rejects_count = "None"

            # Créer un dictionnaire pour stocker les résultats des rejets
            rejects_data = {
                "file": file_path,
                "Numero de BIN": bin_number,
                "total de rejets": transaction_rejects_count,
                "total de xof rejetées": xof_rejects_count
            }

            return rejects_data
        else:
            return {"file": file_path, "error": "ACQUIRER TRANSACTIONS section not found."}

    # Appliquer l'extraction sur les fichiers à travers le dossier des rapports de SETTLEMENT
    rejects_list = []

    #for filename in os.listdir(visa_file_path):
        #if filename.endswith('.TXT'):
            #file_path = os.path.join(visa_file_path, filename)
            #rejects_data = extract_rejects_data(file_path)
            #rejects_list.append(rejects_data)

    # Ouvrir le fichier ZIP et extraire son contenu
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        # Lister tous les fichiers dans le ZIP
        for file_name in zip_ref.namelist():
            if file_name.endswith('.TXT'):
                # Extraire le fichier en mémoire
                with zip_ref.open(file_name) as file:
                    content = file.read().decode('utf-8')
                    # Appliquer l'extraction des données de rejets sur le fichier
                    reject_data = extract_rejects_data(file_name, content)
                    rejects_list.append(reject_data)
    
    # Convertir `rejects_list` en un dictionnaire
    rejects_data_dict_EP100 = {}
    for entry in rejects_list:
        bin_number = entry['Numero de BIN']
        total_rejects = int(entry['total de rejets']) if entry['total de rejets'] != 'None' else 0
        total_xof_rejects = int(entry['total de xof rejetées']) if entry['total de xof rejetées'] != 'None' else 0
        rejects_data_dict_EP100[bin_number] = {
            'Total Rejects': total_rejects, # Nombre de rejets
            'Total Amount': total_xof_rejects  # XOFS rejetées
        }
    

        # Fonction pour extraire les rejets VISA OUTGOING (EP100)
    def extract_EP_rejects(file_path, content):
        filename = os.path.basename(file_path)
        bin_number = filename.split('_')[0]  # Extraire le BIN du nom de fichier

        filiale_name = next((name for name, bin_code in visa_banks_bin.items() if bin_code == bin_number), None)


        # Chercher toutes les sections RECORD dans les fichiers EP100
        record_sections = re.findall(
            r'\s*RECORD\s+----\+----1----\+----2----\+----3----\+----4----\+----5----\+----6----\+----7----\+----8----\+----9----\+---10\s+.*?\s+.*?(.*?)(?=\s*RECORD\s+|$)', 
            content, re.DOTALL
        )

        # Variables pour stocker les ARN et les montants extraits
        arn_list = []
        amount_list = []
        #motif_list = []
        authorization_list = []  # Liste pour stocker les autorisation
        transaction_dates_list = []  # Liste pour stocker les dates de transactions

        
        for section in record_sections:
            # Extraire les ARN de 23 chiffres
            arns = re.findall(r'\d{23}', section)
            amounts = re.findall(r'\d{10}', section)

            # Extraire le numero d'autorisation du fichier EP100A
            authorizations = re.findall(r'1009(D|NO|N)(\S{6})', section)  # Extraire D ou NO ou N suivis par les 6 caracteres


            t_dates = re.findall(r'05070000(0|1)(\d{6})', section)  # trouver la date apres l'expression 05070000 avec 0 ou 1

            # Liste des codes d'autorisations
            authorization_codes = [auth[1] for auth in authorizations]  # auth[1] is the 7-digit number

            transaction_dates = []

            for t_date in t_dates:
                raw_date = t_date[1]  # Extraire la date
                try:
                    # Format de date extraire YYMMDD => nouveau format DD/MM/YYYY
                    formatted_date = datetime.strptime(raw_date, "%y%m%d").strftime("%d/%m/%Y")
                    transaction_dates.append(formatted_date)
                except ValueError:
                    # Si l'extraction est invalide gérer son cas
                    transaction_dates.append('')

            
            authorization_list.extend(authorization_codes)

            transaction_dates_list.extend(transaction_dates)

            motifs = []
            combined_motifs = []

            # On utilise ce bout de code pour utiliser le contenu de fichier en cas de ZIP
            lines = content.splitlines()  # Spliter le contenu par lignes

            for i, line in enumerate(lines):
                # Detecter le motif dans la ligne actuelle
                if re.match(r'^\s*V\d{4}', line):
                    motif = line.strip()
                    
                    # Checker si la ligne suivante contient la suite du motif
                    if i + 1 < len(lines) and lines[i + 1].strip():
                        motif += " " + lines[i + 1].strip()  # Ajouter la ligne suivante au motif
                    
                    
                    motif = re.sub(r'\s+', ' ', motif)  # Remplacer les espaces multiples par un unique espace
                    motifs.append(motif)

            # Vérifier si on a un ARN à la deuxième position et un montant à la septième position (si présents)
            if len(arns) > 1:
                arn_list.append(arns[1])
            if len(amounts) > 6:
                amount_list.append(float(amounts[6]))  # Convertir en float pour calcul précis
            combined_motifs = list(motifs)
        
        # Nombre total de rejets basé sur les ARN
        total_rejects = len(arn_list)

        # Somme totale des montants rejetés
        total_amount = sum(amount_list) if amount_list else 0

        

        # Créer un dictionnaire pour stocker les résultats
        rejects_data = {
            "Filiale": filiale_name,
            #"file": file_path,
            "BIN": bin_number,
            "ARNs": arn_list,
            "Authorization Codes": authorization_list,
            "Transaction dates": transaction_dates_list,
            "Amounts": amount_list,
            "Motifs": combined_motifs,
            "Total Rejects":  total_rejects,  # Somme des rejets totale
            "Total Amount": total_amount # Montant total
        } if arn_list and amount_list else {
            "file": file_path,
            "BIN": bin_number,
            "warning": "No rejected transactions found."
        }
        
        return rejects_data

    # Appliquer l'extraction sur chaque fichier
    #rejects_list = []
    #for filename in os.listdir(visa_rejects_file):
        #if filename.endswith('.TXT'):
            #file_path = os.path.join(visa_rejects_file, filename)
            #rejects_data = extract_EP_rejects(file_path)
            #rejects_list.append(rejects_data)
    with zipfile.ZipFile(zip_reject_path, 'r') as zip_ref:
        # Lister tous les fichiers dans le ZIP
        for file_name in zip_ref.namelist():
            if file_name.endswith('.TXT'):
                # Extraire le fichier en mémoire
                with zip_ref.open(file_name) as file:
                    content = file.read().decode('utf-8')  # Décoder le contenu en UTF-8
                    # Appliquer l'extraction des données de rejet EP sur le fichier
                    rejects_data = extract_EP_rejects(file_name, content)  # Appeler la fonction d'extraction
                    rejects_list.append(rejects_data)
    



    rejects_data_dict = {reject["BIN"]: reject for reject in rejects_list if "Total Rejects" in reject}

    # Fonction pour mettre à jour le nombre total de rejets et le montant de rejets
    def update_rejects_EP100(data, transaction_data, rejects_data_dict):
        for idx, row in data.iterrows():
            filiale = row['FILIALE']
            bin_number = visa_banks_bin.get(filiale)

            # Check if Rapprochement is a string before using 'in'
            rapprochement_status = str(row['Rapprochement']) if not pd.isna(row['Rapprochement']) else ""

            # Vérifier si le numéro de BIN existe dans rejects_data_dict pour les rejets
            if bin_number and bin_number in rejects_data_dict:
                rej_data = rejects_data_dict[bin_number]
                rej_data_100 = rejects_data_dict_EP100[bin_number]
                if 'not ok (avec tr.(s) rejetée(s) a extraire)' in rapprochement_status:
                    # Remplir 'Nbre Total de Rejets' et 'Montants de rejets' les deux
                    data.at[idx, 'Nbre Total de Rejets'] = int(rej_data.get('Total Rejects', 0)) + int(rej_data_100.get('Total Rejects', 0))
                    data.at[idx, 'Montant de Rejets'] = rej_data.get('Total Amount', 0) + rej_data_100.get('Total Amount', 0)
                    
                    
            data['NbreTotaleDeTransactions'] = pd.to_numeric(data['NbreTotaleDeTransactions'], errors='coerce').fillna(0).astype(int)
            data['Nbre Total de Rejets'] = pd.to_numeric(data['Nbre Total de Rejets'], errors='coerce').fillna(0).astype(int)

            data['Montant Total de Transactions'] = pd.to_numeric(data['Montant Total de Transactions'], errors='coerce').fillna(0)
            data['Montant de Rejets'] = pd.to_numeric(data['Montant de Rejets'], errors='coerce').fillna(0)

            # Calculer le nombre de transactions couverts
            data['Nbre de Transactions (Couverture)'] = data.apply(
                lambda row: row['NbreTotaleDeTransactions'] - row['Nbre Total de Rejets'],
                axis=1
            )

            data['Montant de Transactions (Couverture)'] = data.apply(
                    lambda row: row['Montant Total de Transactions'] - row['Montant de Rejets'],
                    axis=1
                )
        

    # Afficher les transactions rejetées pour chaque filiale, en se basant sur le BIN
    for reject in rejects_list:
        print(reject)



    result_df['Rapprochement'] = result_df.apply(compare_transactions, args=(generated_transactions, grouped_filiale), axis=1)
        
    update_rejects(result_df, generated_transactions)

    update_rejects_EP100(result_df, transaction_data, rejects_data_dict)

    # maj des rejets EP100A
    #update_rejects_EP100(data, transaction_data, rejects_data_dict)
      
        
    # Afficher les résultats
    print("reeesult")
    #print(result_df)
    result_df.to_csv("./Reconciliation_Automation_SG/test.csv")


    return result_df, total_nbre_transactions



# Merge the dataframes on relevant common columns
def no_recycled(filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df):

    # Ensure 'TYPE_TRANSACTION' column exists in each DataFrame, if not, create it with default empty values
    if 'TYPE_TRANSACTION' not in filtered_cybersource_df.columns:
        filtered_cybersource_df['TYPE_TRANSACTION'] = ''
    
    if 'TYPE_TRANSACTION' not in filtered_saisie_manuelle_df.columns:
        filtered_saisie_manuelle_df['TYPE_TRANSACTION'] = ''
    
    if 'TYPE_TRANSACTION' not in filtered_pos_df.columns:
        filtered_pos_df['TYPE_TRANSACTION'] = ''

    # Filter cybersource_df to include only rows with TYPE_TRANSACTION == 'ACHAT'
    filtered_cybersource_df = filtered_cybersource_df[filtered_cybersource_df['TYPE_TRANSACTION'] == 'ACHAT']

    # Merge POS and Saisie Manuelle dataframes
    merged_df = pd.merge(
        filtered_pos_df,
        filtered_saisie_manuelle_df,
        on=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION'],
        suffixes=('_pos', '_saisie'),
        how='outer'  # Use outer join to keep all rows from both dataframes
    )

    # Merge with filtered_cybersource_df
    merged_df = pd.merge(
        merged_df,
        filtered_cybersource_df,
        on=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION'],
        suffixes=('_merged', '_cybersource'),
        how='outer'  # Use outer join to keep all rows from all dataframes
    )

    # Fill missing values with 0 and sum the 'NBRE_TRANSACTION' values
    merged_df['NBRE_TRANSACTION'] = (merged_df['NBRE_TRANSACTION_pos'].fillna(0) +
                                     merged_df['NBRE_TRANSACTION_saisie'].fillna(0) +
                                     merged_df['NBRE_TRANSACTION'].fillna(0))
    
    merged_df['NBRE_TRANSACTION'] = merged_df['NBRE_TRANSACTION'].astype(int)

    # Convert 'NBRE_TRANSACTION' to integer
    merged_df['MONTANT_TOTAL'] = (
                                    merged_df['MONTANT_TOTAL_pos'].fillna(0) +
                                    merged_df['MONTANT_TOTAL_saisie'].fillna(0) +
                                    merged_df['MONTANT_TOTAL'].fillna(0)
)
    # Use MONTANT_TOTAL from filtered_pos_df
    #merged_df['MONTANT_TOTAL'] = merged_df['MONTANT_TOTAL_pos'].fillna(0)


    # Use MONTANT_TOTAL from filtered_pos_df
    #merged_df['MONTANT_TOTAL'] = merged_df['MONTANT_TOTAL_pos'].fillna(0)

    # Drop unnecessary columns
    merged_df.drop(['NBRE_TRANSACTION_pos', 'NBRE_TRANSACTION_saisie', 'MONTANT_TOTAL_pos', 'MONTANT_TOTAL_saisie'], axis=1, inplace=True)

    # Drop duplicate rows
    merged_df.drop_duplicates(subset=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION', 'DATE_TRAI'], inplace=True)
    total_nbre_transactions = merged_df['NBRE_TRANSACTION'].sum()

    merged_df = merged_df.reset_index(drop=True)

    return merged_df , total_nbre_transactions


def merging_with_recycled(recycled_rejected_file, filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df, filtering_date, file_path, content, zip_file_path, zip_reject_path):

    result_df, _ = no_recycled(filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df)
    # Lire le fichier des transactions à recycler
    if os.path.exists(recycled_rejected_file):
        # transformer le fichier xlsx en csv puis en df
        df_recycled = excel_to_csv_to_df(recycled_rejected_file)

        # nettoyer les colonnes
        df_recycled.columns = df_recycled.columns.str.strip()
        df_recycled = df_recycled.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

        #Renommer la colonne BANQUE en colonne FILIALE
        df_recycled.rename(columns={'BANQUE': 'FILIALE'}, inplace=True)
        
        # Garder les transactions avec la date retraitement (date de recyclage)
        #df_recycled = standardize_date_format(df_recycled['Date Retraitement'])
        print("date filtereeeeeeeeeeeeeeeeed:")
        print(filtering_date)
        #print("result",df_recycled)
        #df_recycled = df_recycled[df_recycled['Date Retraitement'] == filtering_date]
        df_recycled = df_recycled[df_recycled['Date Retraitement'] == filtering_date.strftime('%Y-%m-%d')]
        #print("date retraitement", df_recycled[df_recycled['Date Retraitement']])
        print(df_recycled)
        #df_recycled['Date Retraitement'] = df_recycled['Date Retraitement'].dt.strftime('%Y-%m-%d')
        df_recycled.drop_duplicates(subset=['FILIALE', 'RESEAU', 'ARN', 'Autorisation', 'Date Transaction', 'Montant', 'Devise'], inplace=True)
        
        # Normaliser le noms de quelques valeurs de colonnes
        df_recycled['FILIALE'] = df_recycled['FILIALE'].str.replace('SG-', 'SG - ')
        df_recycled['FILIALE'] = df_recycled['FILIALE'].str.replace("COTE D'IVOIRE", "COTE D IVOIRE")
        df_recycled['RESEAU'] = df_recycled['RESEAU'].str.replace("VISA", "VISA INTERNATIONAL")


        print("result")
        print(df_recycled)


        # Grouper par FILIALE et RESEAU et calculer et donner le count et la somme du montant
        summary = df_recycled.groupby(['FILIALE', 'RESEAU']).agg(
            NBRE_TRANSACTION=('Montant', 'count'),
            MONTANT_TOTAL=('Montant', 'sum')
        ).reset_index()

        # Printer le resumé des rekets
        print("Résumé des Rejets :")
        st.write("### Le(s) rejet(s) à recycler")
        st.dataframe(summary)
        
        print("result_df_columns", result_df.columns)
        print("recycle columns", df_recycled)
        
        # Fusionner le dataframe des sources fusionnées avec les transactions à recycler
        result_df = result_df.merge(summary, on=['FILIALE', 'RESEAU'], how='left', suffixes=('_merged', '_summary'))

        
        # Remplacer les valeurs NaN par 0
        result_df.fillna(0, inplace=True)
        
        # Initialiser NBRE_TRANSACTION pour eviter les erreurs
        result_df['NBRE_TRANSACTION'] = 0.0

        # préciser la liste des devises sources
        currency_list = ['XOF', 'XAF', 'GNF', 'MGA']

        # Define les conditions de fusionnement
        conditions = [
            (result_df['TYPE_TRANSACTION'] == "ACHAT") & (result_df['CUR'].isin(currency_list)),
            (result_df['TYPE_TRANSACTION'] == "ACHAT") & (result_df['CUR'] == 'EUR'),
            (result_df['TYPE_TRANSACTION'] == "CASH ADVANCE"),
            (result_df['TYPE_TRANSACTION'] == "")
        ]

        # Define Les choix de fusionnement en se basant sur chaque condition
        choices = [
            result_df['NBRE_TRANSACTION_merged'] + result_df['NBRE_TRANSACTION_summary'],  # Pour le type ACHAT (sommer le nombre des transactions rejets recyclées + sources fusionnées)
            result_df['NBRE_TRANSACTION_merged'],   # Pour le type ACHAT avec la devise EUR
            result_df['NBRE_TRANSACTION_merged'],   # Pour le type CASH ADVANCE
            result_df['NBRE_TRANSACTION_merged']    # Pour le type CREDIT VOUCHER
        ]


        # Utiliser np.select pour assigner les valeurs en se basant sur les conditions et les choix
        result_df['NBRE_TRANSACTION'] = np.select(conditions, choices, default=result_df['NBRE_TRANSACTION'])
        
        # S'assurer que les deux colonnes sont numériques avant de sommer
        result_df['MONTANT_TOTAL_merged'] = pd.to_numeric(result_df['MONTANT_TOTAL_merged'], errors='coerce').fillna(0)
        result_df['MONTANT_TOTAL_summary'] = pd.to_numeric(result_df['MONTANT_TOTAL_summary'], errors='coerce').fillna(0)

        # Initialiser MONTANT TOTAL pour eviter les erreurs
        result_df['MONTANT_TOTAL'] = 0.0

        # Define Les choix de fusionnement en se basant sur chaque condition
        choices = [
            result_df['MONTANT_TOTAL_merged'] + result_df['MONTANT_TOTAL_summary'],  # Pour le type ACHAT (sommer les montants)
            result_df['MONTANT_TOTAL_merged'],   # For ACHAT with EUR currency
            result_df['MONTANT_TOTAL_merged'],   # For CASH ADVANCE
            result_df['MONTANT_TOTAL_merged']    # For empty transaction type
        ]

        # Utiliser np.select pour assigner les valeurs en se basant sur les conditions et les choix
        result_df['MONTANT_TOTAL'] = np.select(conditions, choices, default=result_df['MONTANT_TOTAL'])

        # Supprimer les colonnes non necessaires
        result_df.drop(['MONTANT_TOTAL_merged', 'MONTANT_TOTAL_summary'], axis=1, inplace=True)
        result_df.drop(['NBRE_TRANSACTION_summary', 'NBRE_TRANSACTION_merged'], axis=1, inplace=True)

        # Convertir 'NBRE_TRANSACTION' en entier
        result_df['NBRE_TRANSACTION'] = result_df['NBRE_TRANSACTION'].astype(int)

        total_nbre_transactions = result_df['NBRE_TRANSACTION'].sum()
        print(total_nbre_transactions)

        # Dataframe résultat
        result_df = result_df.copy()

        # Définir les colonnes du dataframe du résultat de réconciliation
        new_columns = [
            'FILIALE', 'Réseau', 'Type', 'Date', 'Devise', 'NbreTotaleDeTransactions',
            'Montant Total de Transactions', 'Rapprochement', 'Nbre Total de Rejets',
            'Montant de Rejets', 'Nbre de Transactions (Couverture)', 
            'Montant de Transactions (Couverture)'
        ]

        # Mapper entre les noms de colonnes (result_df et le dataframe résultat)
        column_mapping = {
            'FILIALE': 'FILIALE',
            'RESEAU': 'Réseau',
            'TYPE_TRANSACTION': 'Type',
            'DATE_TRAI': 'Date',
            'CUR': 'Devise',
            'NBRE_TRANSACTION': 'NbreTotaleDeTransactions',
            'MONTANT_TOTAL': 'Montant Total de Transactions'
        }

        # Renommer les colonnes
    result_df.rename(columns=column_mapping, inplace=True)

    # Remplacer les colonnes vides avec colonnes par défaut
    for column in set(new_columns) - set(result_df.columns):
        result_df[column] = ''

    # Réordonner les colonnes
    result_df = result_df[new_columns]

    #print(result_df)

    filename = os.path.basename(file_path)
    bin_number = filename.split('_')[0]  # Extraire BIN des fichiers


     # pour ouvrir en cas de présence de dossier
     # with open(file_path, 'r') as file:
        #content = file.read()
        
    cleaned_content = re.sub(
    r'(ISSUER TRANSACTIONS.*?(?:\*\*\*  END OF VSS-120 REPORT|FINAL SETTLEMENT NET AMOUNT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-130.*?(?:\*\*\*  END OF VSS-130 REPORT)|'
    r'REPORT ID:\s+VSS-140.*?(?:\*\*\*  END OF VSS-140 REPORT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-115.*?(?:\*\*\*  END OF VSS-115 REPORT)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+DISPUTE FIN.*?TOTAL PURCHASE\s+([\d,]+)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+ DISPUTE RESP FIN.*?TOTAL PURCHASE\s+([\d,]+))', 
    '', 
    content, 
    flags=re.DOTALL
)

    # Chercher la section d'extraction commençant de "ACQUIRER TRANSACTIONS" jusqu'à "END OF VSS-120 REPORT"
    acquirer_section = re.search(r'ACQUIRER TRANSACTIONS(.*)(?=\*\*\*  END OF VSS-120 REPORT)', cleaned_content, re.DOTALL)

    if acquirer_section:
        acquirer_content = acquirer_section.group(1).strip()  # Extraire et nettoyer le contenu du bloc

        # Extraire le nombre de transactions TOTAL ORIGINAL SALE
        total_original_sale = re.search(r'TOTAL PURCHASE\s+([\d,]+)', acquirer_content)
        if total_original_sale:
            transaction_count = total_original_sale.group(1).replace(',', '')
        else:
            transaction_count = None

        # Extraire le nombre de transactions MANUAL CASH (CASH ADVANCE)
        manual_cash_strict = re.search(r'^\s*MANUAL CASH\s+(\d+)', acquirer_content, re.MULTILINE)

        total_purchase_mga = re.search(r'CLEARING CURRENCY:\s+MGA.*?TOTAL PURCHASE\s+([\d,]+)', acquirer_content, re.DOTALL)

        # Extract le nombre de transactions TOTAL PURCHASE (ACHAT) pour MGA CLEARING CURRENCY EUR
        nbr_transaction_eur = re.search(r'CLEARING CURRENCY:\s+EUR.*?ORIGINAL SALE\s+(\d+)', acquirer_content, re.DOTALL)
        
        if nbr_transaction_eur:
            transaction_count_EUR = nbr_transaction_eur.group(1).replace(',', '')
        else:
            transaction_count_EUR = None

        if total_purchase_mga:
            transaction_count_MGA = total_purchase_mga.group(1).replace(',', '')
        else:
            transaction_count_MGA = None

        # Extraire le nombre de transactions MERCHANDISE CREDIT (CREDIT VOUCHER)
        merchandise_credit_section = re.search(r'TOTAL MERCHANDISE CREDIT\s+(\d+)', acquirer_content)

        # Extraire le nombre de rejets
        total_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if total_rejects:
            transaction_rejects_count = total_rejects.group(1) #Prendre les elements de la première position comme élement d'extraction
        else:
            transaction_rejects_count = "None"
        
        xof_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if xof_rejects:
            xof_rejects_count = xof_rejects.group(1) #Prendre les elements de la deuxième position comme élement d'extraction (XOFS)
        else:
            xof_rejects_count = "None"

        # Créer un dictionnaire pour stocker les résultats
        transaction_data = {
            "file": file_path,
            "Numero de BIN": bin_number,
            "Nbr de transactions": transaction_count,
            "CASH ADVANCE": None,
            "Nbr de transactions (CLEARING CURR EUR POUR MDG)": None,
            "CREDIT VOUCHER": None,
            "total de rejets" : transaction_rejects_count,
            "total de xof rejetées" : xof_rejects_count
        }

        # Ajouter le nombre de transactions MANUAL CASH (CASH ADVANCE)
        if manual_cash_strict:
            manual_cash_value = manual_cash_strict.group(1)
            if manual_cash_value != "0":  # Ignorer si MANUAL CASH est 0
                transaction_data["CASH ADVANCE"] = manual_cash_value

        # Ajouter le nombre de transactions pour TOTAL PURCHASE (CLEARING CURRENCY EUR) seulement si le BIN est 489316
        if bin_number == "489316":
            transaction_data["Nbr de transactions (CLEARING CURR EUR POUR MDG)"] = transaction_count_EUR
            transaction_data["Nbr de transactions"] = transaction_count_MGA

        # Ajouter le nombre de transactions pour TOTAL MERCHANDISE CREDIT (CREDIT VOUCHER) si ça existe
        if merchandise_credit_section:
            merchandise_credit_count = merchandise_credit_section.group(1)
            transaction_data["CREDIT VOUCHER"] = merchandise_credit_count

    else:
        return {"file": file_path, "error": "ACQUIRER TRANSACTIONS section not found."}
    
    # appliquer l'extraction sur les fichiers à travers le dossier des rapports de SETTLEMENT
    transaction_list = []
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        # Lister tous les fichiers dans le ZIP
        for file_name in zip_ref.namelist():
            if file_name.endswith('.TXT'):
                # Extraire le fichier en mémoire
                with zip_ref.open(file_name) as file:
                    content = file.read().decode('utf-8')
                    # Appliquer l'extraction des données sur le fichier
                    transaction_data = extract_transaction_data(file_name, content)
                    transaction_list.append(transaction_data)
        
    
    # boucler sur le dictionnaire pour afficher les transactions pour chaque filiale (en se basant sur le BIN)
    for transaction in transaction_list:
        print(transaction)


    # Initialiser la mapping des FILIALES à BINs
    visa_banks_bin = {
        'SG - COTE D IVOIRE': '463741',
        'SG - BENIN': '404927',
        'SG - BURKINA FASO': '410282',
        'SG - CAMEROUN': '439972',
        'SG - GUINEE EQUATORIALE': '410655',
        'SG - MADAGASCAR': '489316',
        'SG - SENEGAL': '441358',
        'SG - TCHAD': '458250',
        'SG - CONGO': '464012',
        'SG - GUINEE CONAKRY': '486059'
    }


    # Créer un dictionnaire à partir de la liste des transactions extraites
    generated_transactions = {trans['Numero de BIN']: trans for trans in transaction_list if 'error' not in trans}

    # Pre-calculer (faire la somme) les lignes ayant comme TYPES :  ACHAT et CREDIT VOUCHER groupées par filiale
    grouped_filiale = result_df.groupby('FILIALE').agg(
        achat_transactions=pd.NamedAgg(column='NbreTotaleDeTransactions', aggfunc=lambda x: x[result_df['Type'] == 'ACHAT'].sum()),
        cv_transactions=pd.NamedAgg(column='NbreTotaleDeTransactions', aggfunc=lambda x: x[result_df['Type'].isna() | (result_df['Type'] == '')].sum())
    ).reset_index()

    def safe_float(value):
        """Convert a value to float, handling 'None' and NoneType."""
        if value in (None, 'None'):
            return 0
        try:
            return float(value)
        except ValueError:
            return 0  # Default 0 is la conversion est failed

    def compare_transactions(row, transaction_data, grouped_data):
        filiale = row['FILIALE']
        transaction_type = row['Type']

          
        # Trouver le BIN correspondant par FILIALE
        bin_number = visa_banks_bin.get(filiale)
        if not bin_number:
            return 'Aucun BIN trouvé'
        
        # Extraire les donnèes de transactions par BIN
        trans_data = transaction_data.get(bin_number)
        if not trans_data:
            return 'Aucune donnèe pour BIN'
        
        # Extraire les informations de transactions du dictionnaire
        transaction_count = safe_float(trans_data.get('Nbr de transactions'))
        total_count = safe_float(trans_data.get('Nbr de transactions (CLEARING CURR EUR POUR MDG)'))
        manual_cash = safe_float(trans_data.get('CASH ADVANCE'))
        merchandise_credit = safe_float(trans_data.get('CREDIT VOUCHER'))
        rejects = safe_float(trans_data.get('total de rejets'))
        #xof_rejects = safe_float(trans_data.get('total de xof rejetées'))

        # Get sommes pré-calculées à travers les donnèes groupées
        filiale_data = grouped_data[grouped_data['FILIALE'] == filiale]
        if filiale_data.empty:
            return 'Filiale inconnue'

        achat_transactions = safe_float(filiale_data['achat_transactions'].values[0])
        cv_transactions = safe_float(filiale_data['cv_transactions'].values[0])

        #print(f"FILIALE: {filiale}")
        #print(f"ACHAT transactions (FILIALE): {achat_transactions}")
        #print(f"CREDIT VOUCHER transactions (FILIALE): {cv_transactions}")
        #print(f"Transaction count from settlement report: {transaction_count}")
        #print(f"Rejects from settlement report: {rejects}")
        #print(f"MERCHANDISE CREDIT from settlement report: {merchandise_credit}")

        #Le cas ou on a pas de rejets
        if rejects == 0:
            if transaction_type == 'ACHAT':
                filiale = row['FILIALE']

                # Checker si on a CREDIT VOUCHER (vide/NaN 'Type') pour la meme FILIALE
                cv_rows = result_df[(result_df['FILIALE'] == filiale) & (pd.isna(result_df['Type']) | (result_df['Type'] == ''))]

                # If no CREDIT VOUCHER exists for this FILIALE, apply additional checks
                if cv_rows.empty:
                
                    # Apply checks on transaction counts
                    if row['NbreTotaleDeTransactions'] == transaction_count + merchandise_credit or row['NbreTotaleDeTransactions'] == total_count:
                    
                        return 'ok'
                    else:
                    
                        return 'not ok (avec tr.(s) rejetée(s) a extraire)'

                else:
    
                    if row['NbreTotaleDeTransactions'] == transaction_count or row['NbreTotaleDeTransactions'] == total_count:
                    
                        return 'ok'
                    else:
                    
                        return 'not ok (avec tr.(s) rejetée(s) a extraire)'
                
                
            if transaction_type == 'ACHAT' and transaction_type == '' or  pd.isna(transaction_type):

                # grouper les lignes par filiale
                filiale = row['FILIALE']

                # lignes avec meme filiale pour ACHAT et CREDIT VOUCHER
                relevant_rows = result_df[(result_df['FILIALE'] == filiale) &
                                    ((result_df['Type'] == 'ACHAT') | pd.isna(result_df['Type']) | (result_df['Type'] == ''))]
                

                # Sommer 'NbreTotaleDeTransactions' pour les lignes
                total_transactions = relevant_rows['NbreTotaleDeTransactions'].sum()

                # formule de comparaison
                formule = total_transactions == transaction_count + merchandise_credit
                if formule:
                    return 'ok'
                else:
                    #print(total_transactions)
                    #print(transaction_count + merchandise_credit)
                    return 'not ok (avec tr.(s) rejetée(s) a extraire)'
                
            elif transaction_type == 'CASH ADVANCE':
                if row['NbreTotaleDeTransactions'] == manual_cash:
                    return 'ok'
                else:
                    return 'not ok (avec tr.(s) rejetée(s) a extraire)'
            
        # Le cas ou on a des rejets
        elif rejects != 0:
            try:
                formule_calcul = (achat_transactions + cv_transactions) == (transaction_count - rejects + merchandise_credit)
                print(f"Formule de calcul appliquée ?: {formule_calcul}\n")

                if formule_calcul  and transaction_type == 'ACHAT':
                    return 'not ok'
                elif formule_calcul and (transaction_type == '' or  pd.isna(transaction_type)):
                    return 'ok'
                elif not formule_calcul and (transaction_type == '' or  pd.isna(transaction_type)): 
                    return 'ok'
                elif not formule_calcul and transaction_type == 'ACHAT':
                    return 'not ok (avec tr.(s) rejetée(s) a extraire)'

            except ValueError:
                return 'nok'  # Gérer les erreurs de conversion

        else:
            return 'Type inconnu'
        
    def update_rejects(data, transaction_data):
        for idx, row in data.iterrows():
            # Safely get the 'Rapprochement' value and handle non-string cases
            rapprochement = row.get('Rapprochement', "")
            if not isinstance(rapprochement, str):
                data.at[idx, 'Rapprochement'] = "ok"
                continue
            
            filiale = row['FILIALE']
            bin_number = visa_banks_bin.get(filiale)
            trans_data = transaction_data.get(bin_number)

            if 'not ok (avec tr.(s) rejetée(s) a extraire)' in rapprochement:
                # Fill only 'Nbre Total de Rejets', leave 'Montants de rejets' empty
                if trans_data:
                    data.at[idx, 'Nbre Total de Rejets'] = trans_data.get('total de rejets', 0)
                    data.at[idx, 'Montant de Rejets'] = ""  # Leave it empty as requested
            elif 'not ok' in rapprochement:
                # Fill both 'Nbre Total de Rejets' and 'Montants de rejets'
                if trans_data:
                    data.at[idx, 'Nbre Total de Rejets'] = trans_data.get('total de rejets', 0)
                    data.at[idx, 'Montant de Rejets'] = trans_data.get('total de xof rejetées', 0)
    

        # Fonction pour extraire les rejets des rapports SETTLEMENT de VISA
    def extract_rejects_data(file_path, content):
        filename = os.path.basename(file_path)
        bin_number = filename.split('_')[0]  # Extraire le BIN du nom de fichier

        #with open(file_path, 'r') as file:
            #content = file.read()


        # Nettoyer le contenu pour supprimer les sections non pertinentes
        cleaned_content = re.sub(
        r'(ISSUER TRANSACTIONS.*?(?:\*\*\*  END OF VSS-120 REPORT|FINAL SETTLEMENT NET AMOUNT|TOTAL MERCHANDISE CREDIT)|'
        r'REPORT ID:\s+VSS-130.*?(?:\*\*\*  END OF VSS-130 REPORT)|'
        r'REPORT ID:\s+VSS-140.*?(?:\*\*\*  END OF VSS-140 REPORT|TOTAL MERCHANDISE CREDIT)|'
        r'REPORT ID:\s+VSS-115.*?(?:\*\*\*  END OF VSS-115 REPORT)|'
        r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+DISPUTE FIN.*?TOTAL PURCHASE\s+([\d,]+)|'
        r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+ DISPUTE RESP FIN.*?TOTAL PURCHASE\s+([\d,]+))', 
        '', 
        content, 
        flags=re.DOTALL
    )
        
        # Chercher la section d'extraction commençant de "ACQUIRER TRANSACTIONS" jusqu'à "END OF VSS-120 REPORT"
        acquirer_section = re.search(r'ACQUIRER TRANSACTIONS(.*)(?=\*\*\*  END OF VSS-120 REPORT)', cleaned_content, re.DOTALL)

        if acquirer_section:
            acquirer_content = acquirer_section.group(1).strip()  # Extraire et nettoyer le contenu du bloc

            # Extraire le nombre de rejets
            total_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
            if total_rejects:
                transaction_rejects_count = total_rejects.group(1)  # Prendre les éléments de la première position comme élément d'extraction
            else:
                transaction_rejects_count = "None"
            
            # Extraire le nombre de rejets xof
            xof_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
            if xof_rejects:
                xof_rejects_count = xof_rejects.group(1)  # Prendre les éléments de la deuxième position comme élément d'extraction (XOFS)
            else:
                xof_rejects_count = "None"

            # Créer un dictionnaire pour stocker les résultats des rejets
            rejects_data = {
                "file": file_path,
                "Numero de BIN": bin_number,
                "total de rejets": transaction_rejects_count,
                "total de xof rejetées": xof_rejects_count
            }

            return rejects_data
        else:
            return {"file": file_path, "error": "ACQUIRER TRANSACTIONS section not found."}

    # Appliquer l'extraction sur les fichiers à travers le dossier des rapports de SETTLEMENT
    rejects_list = []

    #for filename in os.listdir(visa_file_path):
        #if filename.endswith('.TXT'):
            #file_path = os.path.join(visa_file_path, filename)
            #rejects_data = extract_rejects_data(file_path)
            #rejects_list.append(rejects_data)

    # Ouvrir le fichier ZIP et extraire son contenu
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        # Lister tous les fichiers dans le ZIP
        for file_name in zip_ref.namelist():
            if file_name.endswith('.TXT'):
                # Extraire le fichier en mémoire
                with zip_ref.open(file_name) as file:
                    content = file.read().decode('utf-8')
                    # Appliquer l'extraction des données de rejets sur le fichier
                    reject_data = extract_rejects_data(file_name, content)
                    rejects_list.append(reject_data)
    
    # Convertir `rejects_list` en un dictionnaire
    rejects_data_dict_EP100 = {}
    for entry in rejects_list:
        bin_number = entry['Numero de BIN']
        total_rejects = int(entry['total de rejets']) if entry['total de rejets'] != 'None' else 0
        total_xof_rejects = int(entry['total de xof rejetées']) if entry['total de xof rejetées'] != 'None' else 0
        rejects_data_dict_EP100[bin_number] = {
            'Total Rejects': total_rejects, # Nombre de rejets
            'Total Amount': total_xof_rejects  # XOFS rejetées
        }
    

        # Fonction pour extraire les rejets VISA OUTGOING (EP100)
    def extract_EP_rejects(file_path, content):
        filename = os.path.basename(file_path)
        bin_number = filename.split('_')[0]  # Extraire le BIN du nom de fichier

        filiale_name = next((name for name, bin_code in visa_banks_bin.items() if bin_code == bin_number), None)


        # Chercher toutes les sections RECORD dans les fichiers EP100
        record_sections = re.findall(
            r'\s*RECORD\s+----\+----1----\+----2----\+----3----\+----4----\+----5----\+----6----\+----7----\+----8----\+----9----\+---10\s+.*?\s+.*?(.*?)(?=\s*RECORD\s+|$)', 
            content, re.DOTALL
        )

        # Variables pour stocker les ARN et les montants extraits
        arn_list = []
        amount_list = []
        #motif_list = []
        authorization_list = []  # Liste pour stocker les autorisation
        transaction_dates_list = []  # Liste pour stocker les dates de transactions

        
        for section in record_sections:
            # Extraire les ARN de 23 chiffres
            arns = re.findall(r'\d{23}', section)
            amounts = re.findall(r'\d{10}', section)

            # Extraire le numero d'autorisation du fichier EP100A
            authorizations = re.findall(r'1009(D|NO|N)(\S{6})', section)  # Extraire D ou NO ou N suivis par les 6 caracteres


            t_dates = re.findall(r'05070000(0|1)(\d{6})', section)  # trouver la date apres l'expression 05070000 avec 0 ou 1

            # Liste des codes d'autorisations
            authorization_codes = [auth[1] for auth in authorizations]  # auth[1] is the 7-digit number

            transaction_dates = []

            for t_date in t_dates:
                raw_date = t_date[1]  # Extraire la date
                try:
                    # Format de date extraire YYMMDD => nouveau format DD/MM/YYYY
                    formatted_date = datetime.strptime(raw_date, "%y%m%d").strftime("%d/%m/%Y")
                    transaction_dates.append(formatted_date)
                except ValueError:
                    # Si l'extraction est invalide gérer son cas
                    transaction_dates.append('')

            
            authorization_list.extend(authorization_codes)

            transaction_dates_list.extend(transaction_dates)

            motifs = []
            combined_motifs = []

            # On utilise ce bout de code pour utiliser le contenu de fichier en cas de ZIP
            lines = content.splitlines()  # Spliter le contenu par lignes

            for i, line in enumerate(lines):
                # Detecter le motif dans la ligne actuelle
                if re.match(r'^\s*V\d{4}', line):
                    motif = line.strip()
                    
                    # Checker si la ligne suivante contient la suite du motif
                    if i + 1 < len(lines) and lines[i + 1].strip():
                        motif += " " + lines[i + 1].strip()  # Ajouter la ligne suivante au motif
                    
                    
                    motif = re.sub(r'\s+', ' ', motif)  # Remplacer les espaces multiples par un unique espace
                    motifs.append(motif)

            # Vérifier si on a un ARN à la deuxième position et un montant à la septième position (si présents)
            if len(arns) > 1:
                arn_list.append(arns[1])
            if len(amounts) > 6:
                amount_list.append(float(amounts[6]))  # Convertir en float pour calcul précis
            combined_motifs = list(motifs)
        
        # Nombre total de rejets basé sur les ARN
        total_rejects = len(arn_list)

        # Somme totale des montants rejetés
        total_amount = sum(amount_list) if amount_list else 0

        

        # Créer un dictionnaire pour stocker les résultats
        rejects_data = {
            "Filiale": filiale_name,
            #"file": file_path,
            "BIN": bin_number,
            "ARNs": arn_list,
            "Authorization Codes": authorization_list,
            "Transaction dates": transaction_dates_list,
            "Amounts": amount_list,
            "Motifs": combined_motifs,
            "Total Rejects":  total_rejects,  # Somme des rejets totale
            "Total Amount": total_amount # Montant total
        } if arn_list and amount_list else {
            "file": file_path,
            "BIN": bin_number,
            "warning": "No rejected transactions found."
        }
        
        return rejects_data

    # Appliquer l'extraction sur chaque fichier
    #rejects_list = []
    #for filename in os.listdir(visa_rejects_file):
        #if filename.endswith('.TXT'):
            #file_path = os.path.join(visa_rejects_file, filename)
            #rejects_data = extract_EP_rejects(file_path)
            #rejects_list.append(rejects_data)
    with zipfile.ZipFile(zip_reject_path, 'r') as zip_ref:
        # Lister tous les fichiers dans le ZIP
        for file_name in zip_ref.namelist():
            if file_name.endswith('.TXT'):
                # Extraire le fichier en mémoire
                with zip_ref.open(file_name) as file:
                    content = file.read().decode('utf-8')  # Décoder le contenu en UTF-8
                    # Appliquer l'extraction des données de rejet EP sur le fichier
                    rejects_data = extract_EP_rejects(file_name, content)  # Appeler la fonction d'extraction
                    rejects_list.append(rejects_data)
    



    rejects_data_dict = {reject["BIN"]: reject for reject in rejects_list if "Total Rejects" in reject}

    # Fonction pour mettre à jour le nombre total de rejets et le montant de rejets
    def update_rejects_EP100(data, transaction_data, rejects_data_dict):
        for idx, row in data.iterrows():
            filiale = row['FILIALE']
            bin_number = visa_banks_bin.get(filiale)

            #if row['Type'] == "ACHAT" and row['Nbre Total de Rejets'] == 0:
                #data.at[idx, 'Rapprochement'] = 'ok'

            # Check if Rapprochement is a string before using 'in'
            rapprochement_status = str(row['Rapprochement']) if not pd.isna(row['Rapprochement']) else ""

            # Vérifier si le numéro de BIN existe dans rejects_data_dict pour les rejets
            if bin_number and bin_number in rejects_data_dict:
                rej_data = rejects_data_dict[bin_number]
                rej_data_100 = rejects_data_dict_EP100[bin_number]
                if 'not ok (avec tr.(s) rejetée(s) a extraire)' in rapprochement_status:
                    # Remplir 'Nbre Total de Rejets' et 'Montants de rejets' les deux
                    data.at[idx, 'Nbre Total de Rejets'] = int(rej_data.get('Total Rejects', 0)) + int(rej_data_100.get('Total Rejects', 0))
                    data.at[idx, 'Montant de Rejets'] = rej_data.get('Total Amount', 0) + rej_data_100.get('Total Amount', 0)
                    
                    
            data['NbreTotaleDeTransactions'] = pd.to_numeric(data['NbreTotaleDeTransactions'], errors='coerce').fillna(0).astype(int)
            data['Nbre Total de Rejets'] = pd.to_numeric(data['Nbre Total de Rejets'], errors='coerce').fillna(0).astype(int)

            data['Montant Total de Transactions'] = pd.to_numeric(data['Montant Total de Transactions'], errors='coerce').fillna(0)
            data['Montant de Rejets'] = pd.to_numeric(data['Montant de Rejets'], errors='coerce').fillna(0)

            # Calculer le nombre de transactions couverts
            data['Nbre de Transactions (Couverture)'] = data.apply(
                lambda row: row['NbreTotaleDeTransactions'] - row['Nbre Total de Rejets'],
                axis=1
            )

            data['Montant de Transactions (Couverture)'] = data.apply(
                    lambda row: row['Montant Total de Transactions'] - row['Montant de Rejets'],
                    axis=1
                )

            
    # Afficher les transactions rejetées pour chaque filiale, en se basant sur le BIN
    for reject in rejects_list:
        print(reject)



    result_df['Rapprochement'] = result_df.apply(compare_transactions, args=(generated_transactions, grouped_filiale), axis=1)
        
    update_rejects(result_df, generated_transactions)

    update_rejects_EP100(result_df, transaction_data, rejects_data_dict)

    # maj des rejets EP100A
    #update_rejects_EP100(data, transaction_data, rejects_data_dict)
      
        
    # Afficher les résultats
    print("reeesult")
    #print(result_df)
    result_df.to_csv("./Reconciliation_Automation_SG/test.csv")


    return result_df


# Initialiser la mapping des FILIALES à BINs
visa_banks_bin = {
    'SG - COTE D IVOIRE': '463741',
    'SG - BENIN': '404927',
    'SG - BURKINA FASO': '410282',
    'SG - CAMEROUN': '439972',
    'SG - GUINEE EQUATORIALE': '410655',
    'SG - MADAGASCAR': '489316',
    'SG - SENEGAL': '441358',
    'SG - TCHAD': '458250',
    'SG - CONGO': '464012',
    'SG - GUINEE CONAKRY': '486059'
}


def get_filiale_from_bin(bin_number):
    """
    Trouver la filiale correspondant au numéro de BIN.
    """
    for filiale, bin_value in visa_banks_bin.items():
        if bin_number == bin_value:
            return filiale
    return "Unknown FILIALE"

# Fonction pour extraire le nombres de transactions des rapports SETTLEMENT de VISA
def extract_transaction_data(file_path, content):
    filename = os.path.basename(file_path)
    bin_number = filename.split('_')[0]  # Extraire BIN des fichiers


     # pour ouvrir en cas de présence de dossier
     # with open(file_path, 'r') as file:
        #content = file.read()
        
    cleaned_content = re.sub(
    r'(ISSUER TRANSACTIONS.*?(?:\*\*\*  END OF VSS-120 REPORT|FINAL SETTLEMENT NET AMOUNT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-130.*?(?:\*\*\*  END OF VSS-130 REPORT)|'
    r'REPORT ID:\s+VSS-140.*?(?:\*\*\*  END OF VSS-140 REPORT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-115.*?(?:\*\*\*  END OF VSS-115 REPORT)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+DISPUTE FIN.*?TOTAL PURCHASE\s+([\d,]+)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+ DISPUTE RESP FIN.*?TOTAL PURCHASE\s+([\d,]+))', 
    '', 
    content, 
    flags=re.DOTALL
)

    # Chercher la section d'extraction commençant de "ACQUIRER TRANSACTIONS" jusqu'à "END OF VSS-120 REPORT"
    acquirer_section = re.search(r'ACQUIRER TRANSACTIONS(.*)(?=\*\*\*  END OF VSS-120 REPORT)', cleaned_content, re.DOTALL)

    if acquirer_section:
        acquirer_content = acquirer_section.group(1).strip()  # Extraire et nettoyer le contenu du bloc

        # Extraire le nombre de transactions TOTAL ORIGINAL SALE
        total_original_sale = re.search(r'TOTAL PURCHASE\s+([\d,]+)', acquirer_content)
        if total_original_sale:
            transaction_count = total_original_sale.group(1).replace(',', '')
        else:
            transaction_count = None

        # Extraire le nombre de transactions MANUAL CASH (CASH ADVANCE)
        manual_cash_strict = re.search(r'^\s*MANUAL CASH\s+(\d+)', acquirer_content, re.MULTILINE)

        total_purchase_mga = re.search(r'CLEARING CURRENCY:\s+MGA.*?TOTAL PURCHASE\s+([\d,]+)', acquirer_content, re.DOTALL)

        # Extract le nombre de transactions TOTAL PURCHASE (ACHAT) pour MGA CLEARING CURRENCY EUR
        nbr_transaction_eur = re.search(r'CLEARING CURRENCY:\s+EUR.*?ORIGINAL SALE\s+(\d+)', acquirer_content, re.DOTALL)
        
        if nbr_transaction_eur:
            transaction_count_EUR = nbr_transaction_eur.group(1).replace(',', '')
        else:
            transaction_count_EUR = None

        if total_purchase_mga:
            transaction_count_MGA = total_purchase_mga.group(1).replace(',', '')
        else:
            transaction_count_MGA = None

        # Extraire le nombre de transactions MERCHANDISE CREDIT (CREDIT VOUCHER)
        merchandise_credit_section = re.search(r'TOTAL MERCHANDISE CREDIT\s+(\d+)', acquirer_content)

        # Extraire le nombre de rejets
        total_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if total_rejects:
            transaction_rejects_count = total_rejects.group(1) #Prendre les elements de la première position comme élement d'extraction
        else:
            transaction_rejects_count = "None"
        
        xof_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if xof_rejects:
            xof_rejects_count = xof_rejects.group(1) #Prendre les elements de la deuxième position comme élement d'extraction (XOFS)
        else:
            xof_rejects_count = "None"

        # Créer un dictionnaire pour stocker les résultats
        transaction_data = {
            #"file": file_path,
            "FILIALE": get_filiale_from_bin(bin_number),  # Ajouter la filiale correspondante
            "Numero de BIN": bin_number,
            "Nbr de transactions": transaction_count,
            "CASH ADVANCE": None,
            "Nbr de transactions (CLEARING CURR EUR POUR MDG)": None,
            "CREDIT VOUCHER": None,
            "total de rejets" : transaction_rejects_count,
            "total de xof rejetées" : xof_rejects_count
        }

        # Ajouter le nombre de transactions MANUAL CASH (CASH ADVANCE)
        if manual_cash_strict:
            manual_cash_value = manual_cash_strict.group(1)
            if manual_cash_value != "0":  # Ignorer si MANUAL CASH est 0
                transaction_data["CASH ADVANCE"] = manual_cash_value

        # Ajouter le nombre de transactions pour TOTAL PURCHASE (CLEARING CURRENCY EUR) seulement si le BIN est 489316
        if bin_number == "489316":
            transaction_data["Nbr de transactions (CLEARING CURR EUR POUR MDG)"] = transaction_count_EUR
            transaction_data["Nbr de transactions"] = transaction_count_MGA

        # Ajouter le nombre de transactions pour TOTAL MERCHANDISE CREDIT (CREDIT VOUCHER) si ça existe
        if merchandise_credit_section:
            merchandise_credit_count = merchandise_credit_section.group(1)
            transaction_data["CREDIT VOUCHER"] = merchandise_credit_count

        return transaction_data
    else:
        return {"file": file_path, "error": "ACQUIRER TRANSACTIONS section not found."}
    


# Fonction pour extraire les rejets des rapports SETTLEMENT de VISA
def extract_rejects_data(file_path, content):
    filename = os.path.basename(file_path)
    bin_number = filename.split('_')[0]  # Extraire le BIN du nom de fichier

    #with open(file_path, 'r') as file:
        #content = file.read()


    # Nettoyer le contenu pour supprimer les sections non pertinentes
    cleaned_content = re.sub(
    r'(ISSUER TRANSACTIONS.*?(?:\*\*\*  END OF VSS-120 REPORT|FINAL SETTLEMENT NET AMOUNT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-130.*?(?:\*\*\*  END OF VSS-130 REPORT)|'
    r'REPORT ID:\s+VSS-140.*?(?:\*\*\*  END OF VSS-140 REPORT|TOTAL MERCHANDISE CREDIT)|'
    r'REPORT ID:\s+VSS-115.*?(?:\*\*\*  END OF VSS-115 REPORT)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+DISPUTE FIN.*?TOTAL PURCHASE\s+([\d,]+)|'
    r'ACQUIRER TRANSACTIONS\s+PURCHASE\s+ DISPUTE RESP FIN.*?TOTAL PURCHASE\s+([\d,]+))', 
    '', 
    content, 
    flags=re.DOTALL
)
    
    # Chercher la section d'extraction commençant de "ACQUIRER TRANSACTIONS" jusqu'à "END OF VSS-120 REPORT"
    acquirer_section = re.search(r'ACQUIRER TRANSACTIONS(.*)(?=\*\*\*  END OF VSS-120 REPORT)', cleaned_content, re.DOTALL)

    if acquirer_section:
        acquirer_content = acquirer_section.group(1).strip()  # Extraire et nettoyer le contenu du bloc

        # Extraire le nombre de rejets
        total_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if total_rejects:
            transaction_rejects_count = total_rejects.group(1)  # Prendre les éléments de la première position comme élément d'extraction
        else:
            transaction_rejects_count = "None"
        
        # Extraire le nombre de rejets xof
        xof_rejects = re.search(r'ORIGINAL SALE\s+RETURN\s+[A-Z0-9]+\s+(\d+)', acquirer_content)
        if xof_rejects:
            xof_rejects_count = xof_rejects.group(1)  # Prendre les éléments de la deuxième position comme élément d'extraction (XOFS)
        else:
            xof_rejects_count = "None"

        # Créer un dictionnaire pour stocker les résultats des rejets
        rejects_data = {
            "file": file_path,
            "Numero de BIN": bin_number,
            "total de rejets": transaction_rejects_count,
            "total de xof rejetées": xof_rejects_count
        }

        rejects_list = []
        rejects_list.append(rejects_data)
        rejects_data_dict_EP100 = {}
        for entry in rejects_list:
            bin_number = entry['Numero de BIN']
            total_rejects = int(entry['total de rejets']) if entry['total de rejets'] != 'None' else 0
            total_xof_rejects = int(entry['total de xof rejetées']) if entry['total de xof rejetées'] != 'None' else 0
            rejects_data_dict_EP100[bin_number] = {
                'Total Rejects': total_rejects, # Nombre de rejets
                'Total Amount': total_xof_rejects  # XOFS rejetées
            }
        return rejects_list
    else:
        return {"file": file_path, "error": "ACQUIRER TRANSACTIONS section not found."}
    




# Fonction pour extraire les rejets VISA OUTGOING (EP100)
def extract_EP_rejects(file_path, content):
    filename = os.path.basename(file_path)
    bin_number = filename.split('_')[0]  # Extract BIN from filename

    filiale_name = next((name for name, bin_code in visa_banks_bin.items() if bin_code == bin_number), None)

    # Find all RECORD sections in the EP100 files
    record_sections = re.findall(
        r'\s*RECORD\s+----\+----1----\+----2----\+----3----\+----4----\+----5----\+----6----\+----7----\+----8----\+----9----\+---10\s+.*?\s+.*?(.*?)(?=\s*RECORD\s+|$)', 
        content, re.DOTALL
    )

    # Variables to store ARNs and extracted amounts
    arn_list = []
    amount_list = []
    authorization_list = []  # List to store authorization codes
    transaction_dates_list = []  # List to store transaction dates

    # Initialize motifs and combined_motifs outside the loop
    motifs = []
    combined_motifs = []

     # Variables pour stocker les ARN et les montants extraits
    arn_list = []
    amount_list = []
    #motif_list = []
    authorization_list = []  # Liste pour stocker les autorisation
    transaction_dates_list = []  # Liste pour stocker les dates de transactions

    
    for section in record_sections:
        # Extraire les ARN de 23 chiffres
        arns = re.findall(r'\d{23}', section)
        amounts = re.findall(r'\d{10}', section)

        # Extraire le numero d'autorisation du fichier EP100A
        authorizations = re.findall(r'1009(D|NO|N)(\S{6})', section)  # Extraire D ou NO ou N suivis par les 6 caracteres


        t_dates = re.findall(r'05070000(0|1)(\d{6})', section)  # trouver la date apres l'expression 05070000 avec 0 ou 1

        # Liste des codes d'autorisations
        authorization_codes = [auth[1] for auth in authorizations]  # auth[1] is the 7-digit number

        transaction_dates = []

        for t_date in t_dates:
            raw_date = t_date[1]  # Extraire la date
            try:
                # Format de date extraire YYMMDD => nouveau format DD/MM/YYYY
                formatted_date = datetime.strptime(raw_date, "%y%m%d").strftime("%d/%m/%Y")
                transaction_dates.append(formatted_date)
            except ValueError:
                # Si l'extraction est invalide gérer son cas
                transaction_dates.append('')

        
        authorization_list.extend(authorization_codes)

        transaction_dates_list.extend(transaction_dates)

        motifs = []
        combined_motifs = []

        # On utilise ce bout de code pour utiliser le contenu de fichier en cas de ZIP
        lines = content.splitlines()  # Spliter le contenu par lignes

        for i, line in enumerate(lines):
            # Detecter le motif dans la ligne actuelle
            if re.match(r'^\s*V\d{4}', line):
                motif = line.strip()
                
                # Checker si la ligne suivante contient la suite du motif
                if i + 1 < len(lines) and lines[i + 1].strip():
                    motif += " " + lines[i + 1].strip()  # Ajouter la ligne suivante au motif
                
                
                motif = re.sub(r'\s+', ' ', motif)  # Remplacer les espaces multiples par un unique espace
                motifs.append(motif)

        # Vérifier si on a un ARN à la deuxième position et un montant à la septième position (si présents)
        if len(arns) > 1:
            arn_list.append(arns[1])
        if len(amounts) > 6:
            amount_list.append(float(amounts[6]))  # Convertir en float pour calcul précis
        combined_motifs = list(motifs)
    
    # Nombre total de rejets basé sur les ARN
    total_rejects = len(arn_list)

    # Somme totale des montants rejetés
    total_amount = sum(amount_list) if amount_list else 0

    # Create a dictionary to store results
    rejects_data = {
        "Filiale": filiale_name,
        #"file": file_path,
        #"BIN": bin_number,
        "ARNs": arn_list,
        "Authorization Codes": authorization_list,
        "Transaction dates": transaction_dates_list,
        "Amounts": amount_list,
        "Motifs": combined_motifs,
        "Total Rejects": total_rejects,  # Total number of rejects
        "Total Amount": total_amount  # Total amount rejected
    }

    # Exporter les résultats dans un format CSV
    data_for_export = []

    # Create lists directly from rejects_data
    filiale = rejects_data.get("Filiale", "")
    arns = rejects_data.get("ARNs", [])
    authorization_codes = rejects_data.get("Authorization Codes", [])
    transaction_dates = rejects_data.get("Transaction dates", [])
    amounts = rejects_data.get("Amounts", [])
    motifs = rejects_data.get("Motifs", [])

    # Ensure the lists have the same length by filling missing elements with empty strings
    max_length = max(len(arns), len(amounts), len(motifs))
    arns.extend([""] * (max_length - len(arns)))
    authorization_codes.extend([""] * (max_length - len(authorization_codes)))
    transaction_dates.extend([""] * (max_length - len(transaction_dates)))
    #currency_codes.extend([""] * (max_length - len(currency_codes)))
    amounts.extend([""] * (max_length - len(amounts)))
    motifs.extend([""] * (max_length - len(motifs)))


    for arn, amount, authorization_code, transaction_date, motif in zip(arns, amounts, authorization_codes, transaction_dates, motifs):
        rejectes_transactions = {
            "Filiale": filiale,
            "ARN": arn,
            "Authorization code": authorization_code,
            "Transaction Date": transaction_date,
            "Amount": amount,
            "Motif": motif,
        }
        data_for_export.append(rejectes_transactions)

    # Create a DataFrame and export
    df =  pd.DataFrame(data_for_export)

    # Afficher les résultats
    print("data", df)


    return df



def blue_style_and_save_to_excel(df):
    """
    Styles a DataFrame and saves it to an Excel file with a predefined style.

    Parameters:
    - df (pd.DataFrame): The DataFrame to be styled and saved.

    Returns:
    - str: Path to the saved Excel file.
    """
    # Define the path for the output Excel file
    excel_path = './styled_data.xlsx'

    # Save DataFrame to an Excel file
    df.to_excel(excel_path, index=False)

    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    # Define the table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=True
    )

    # Add the table to the worksheet
    tab = Table(displayName="Table1", ref=sheet.dimensions)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    # Define the number format for thousands and decimal separators
    number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Apply additional styling
    # Set column widths based on content length and apply number format
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            # Apply number format to columns containing 'Montant'
            if 'Montant' in df.columns[col[0].column - 1]:
                cell.number_format = number_format
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Set the header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the styled workbook
    workbook.save(excel_path)

    return excel_path

def styling_and_saving_reconciliated(excel_path):
    """
    Styles specific cells in an existing Excel file based on 'Rapprochement' column values.

    Parameters:
    - excel_path (str): Path to the Excel file to be styled.

    Returns:
    - str: Path to the styled Excel file.
    """
    # Reopen the workbook with openpyxl to apply styles
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['Sheet1']

    # Load DataFrame from the Excel file
    df = pd.read_excel(excel_path, sheet_name='Sheet1')

    # Apply styles to specific cells
    for row_idx, row in df.iterrows():
        for col_idx in range(len(row)):
            cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)  # +2 to account for header and 1-based index
            # Apply bold font to 'Rapprochement' column cells
            if col_idx == row.index.get_loc('Rapprochement'):
                cell.font = Font(bold=True)
            # Apply red background for 'not ok' and white text color
            if row['Rapprochement'] == 'not ok' or row['Rapprochement'] == 'not ok (avec tr.(s) rejetée(s) a extraire)':
                cell.fill = PatternFill(start_color='ffe26b0a', end_color='ffe26b0a', fill_type="solid")  # Red
                cell.font = Font(bold=True ,color="FFFFFF")  # Set text color to white

    # Save the styled workbook
    workbook.save(excel_path)
    return excel_path


def highlight_non_reconciliated_row(row):
    return [f'background-color: #ffab77; font-weight: bold;'
            if (row['Rapprochement'] == 'not ok' or row['Rapprochement'] == 'not ok (avec tr.(s) rejetée(s) a extraire)') and row['Type'] == 'ACHAT' and row['Devise'] != 'EUR'  else '' for _ in row]

def download_file(recon, df, file_partial_name, button_label , run_date):
    # Assuming styling_and_saving_reconciliated is a defined function that processes the DataFrame
    excel_path1 = blue_style_and_save_to_excel(df)

    if recon :
        excel_path2 = styling_and_saving_reconciliated(excel_path1)
    else:
        excel_path2 = excel_path1

    with open(excel_path2, 'rb') as f:
        excel_data = io.BytesIO(f.read())


    with open(excel_path2, 'rb') as f:
        excel_data = io.BytesIO(f.read())

    # Define the file name
    file_name = f"{file_partial_name}.xlsx"

    # Create a download button for the Excel file
    st.download_button(
        label=button_label,
        data=excel_data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )
    return excel_path2 , file_name
    
import tempfile
def save_excel_locally(excel_path , file_name):
    wb = openpyxl.load_workbook(excel_path)

    # Save the workbook locally with original file name
    temp_dir = tempfile.gettempdir()  # Get the temporary directory
    file_path = os.path.join(temp_dir, file_name)  # Define the file path

    wb.save(file_path)  # Save the workbook

    return file_path  # Return the file path







